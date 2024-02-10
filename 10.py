import shutil
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Combobox
from typing import List, Union, Any
import win32com.client as win32
import openpyxl

window = tk.Tk()
window.title("Перечетная ведомость")
window.geometry('1005x250')

with open('addresses.txt','r',encoding='utf-8') as source:
    adressess = source.read().split('\n')

ages:list[str] = ["Более 10", "Менее 10"]
conclusion:list[str] = ["Вырубить", "Сохранить", "Пересадить"]
species:list[str] = ['Акация белая', 'акация желтая куст', 'барбарис куст', 'Бархат амурский', 'Береза', 'бересклет куст',
           'бирючина куст', 'Боярышник', 'боярышник куст', 'бузина куст', 'Вишня', 'вишня куст', 'Вяз', 'Газон',
           'Груша', 'дек.-лиственный кустарник', 'дерен куст', 'Дуб', 'Ель', 'жимолость куст', 'Ива', 'ива (поросль)',
           'Ива белая', 'ирга куст', 'калина куст', 'Каштан', 'кизильник куст', 'Клен', 'клен куст',
           'Клен ясенелистный', 'клен ясенелистный (поросль)', 'крушина куст', 'кустарник разный', 'лещина куст',
           'Лжетсуга', 'лиана куст', 'Липа', 'Лиственница', 'можжевельник куст', 'Ольха', 'Орех', 'Осина', 'Остолоп',
           'Пень', 'Пихта', 'пл.-ягодный кустарник', 'Плодовое', 'поросль', 'пузыреплодник куст', 'ракитник куст',
           'роза куст', 'Рябина', 'Саженцы', 'Самосев до 8 см.', 'сирень куст', 'Слива', 'слива куст', 'смородина куст',
           'снежноягодник куст', 'Сосна', 'спирея куст', 'Сухостой', 'Тополь', 'тополь (поросль)', 'Тополь белый',
           'Тополь пирамидальный', 'Травяной покров', 'Туя', 'туя куст', 'хвойный кустарник', 'Черемуха',
           'черемуха куст', 'чубушник куст', 'Экзот', 'Яблоня', 'Ясень']

def suggest_adress(event)->None:
    """
    suggests address autoincrement in search field combobox
    :type event: object

    """
    value = event.widget.get()
    if value == '':
        ad_value['values'] = adressess
    else :
        data: list[Union[str, Any]] = []
        for item in adressess:
            if value.lower() in item.lower():
                data.append(item)
        ad_value['values'] = data


Label(window, text="Адрес, начните вводить значение, затем нажмите стрелку:").grid(row=0, column=0, sticky=W)
ad_value: Combobox = ttk.Combobox(window, value=adressess, width = 50)
ad_value.bind('<KeyRelease>', suggest_adress)
ad_value.grid(column=0, row=1)


def suggest_species(event)->None:
    """
    suggests species autoincrement in search field combobox
    :type event: object

    """
    value = event.widget.get()
    if value == '':
        ad_value2['values'] = species
    else :
        data: list[str] = []
        for item in species:
            if value.lower() in item.lower():
                data.append(item)
        ad_value2['values'] = data


Label(window, text="Вид, начните вводить значение в поле:").grid(row=2, column=0, sticky=W)        
ad_value2:Combobox = ttk.Combobox(window, value=species, width = 50)
ad_value2.bind('<KeyRelease>', suggest_species)
ad_value2.grid(column=0, row=3)
Label(window, text="Возраст:").grid(row=1, column=2, sticky=W)
ad_value9:Combobox = ttk.Combobox(window, value = ages, width = 15)
ad_value9.grid(column=2, row=2)


def edit_cell ()->None:
    """
    function allows change cells in xlsx file
    """
    wb = openpyxl.load_workbook('1.xlsx')
    ws = wb['1']
    ws['F9'].value = ad_value9.get()
    ws['G9'].value = ad_value10.get()
    ws['I9'].value = ad_value11.get()
    ws['A2'].value = str("Перечетная ведомость деревьев и кустарников по адресу: г. Москва, ЮЗАО,") + ad_value.get()
    ws['C9'].value = ws['C10'].value = ws['C12'].value = ad_value3.get()
    ws['C13'].value = ws['C10'].value = ws['C9'].value = ad_value4.get()
    ws['C22'].value = ad_value5.get()
    ws['B9'].value = ad_value2.get()
    ws['D29'].value = 'Дата ' + ad_value12.get() + ' г.'
    ws['C23'].value = ad_value6.get()
    ws['C24'].value = ad_value7.get()
    ws['E9'].value = ad_value8.get()
    ws['H9'].value = ad_value13.get()
    wb.save('1.xlsx')

    #Convertise xlsx to pdf
    shutil.copyfile('1.xlsx', 'C:/425/felling_ticket/2.xlsx')
    
    excel_file:str = 'C:/425/felling_ticket/2.xlsx'
    pdf_file:str = '2.pdf'

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(excel_file)

    wb.SaveAs(pdf_file, FileFormat=57)
    
    for img in ws._images:
    # Получаем свойства картинки
        _, _, _, _, cell = img.anchor
        # Удаляем картинку из ячейки
        ws._images.remove(img)
        # Удаляем свойства картинки из ячейки
        del ws._drawing[str(cell)]
    wb.save('1.xlsx')
    wb.Close()
    excel.Quit()

Button(window, text="Ввести", command=lambda: edit_cell()).grid(row=3, column=5, sticky=W)
Label(window, text="Высота, м:").grid(row=1, column=3, sticky=W)
ad_value10:Entry = ttk.Entry(width = 5)
ad_value10.grid(column=3, row=2)
Label(window, text="Сохранить, кол-во:").grid(row=5, column=0, sticky=W)
ad_value3:Entry = ttk.Entry(width = 5)
ad_value3.grid(column=0, row=5)
Label(window, text="Вырубить, кол-во:").grid(row=6, column=0, sticky=W)
ad_value4:Entry = ttk.Entry(width = 5)
ad_value4.grid(column=0, row=6)
Label(window, text="Аварийное, кол-во:").grid(row=7, column=0, sticky=W)
ad_value5:Entry = ttk.Entry(width = 5)
ad_value5.grid(column=0, row=7)
Label(window, text="Заключение:").grid(row=1, column=5, sticky=W)
ad_value11:Combobox = ttk.Combobox(window, value = conclusion, width = 15)
ad_value11.grid(column=5, row=2)
Label(window, text="Число, месяц, год:").grid(row=10, column=1, sticky=W)
ad_value12:Entry = ttk.Entry(width = 10)
ad_value12.grid(column=1, row=11)
Label(window, text="Характеристика состояния зеленых насаждений:").grid(row=1, column=4, sticky=W)
ad_value13:Entry = ttk.Entry(width = 25)
ad_value13.grid(column=4, row=2)
Label(window, text="сухостой, кол-во:").grid(row=8, column=0, sticky=W)
ad_value6:Entry = ttk.Entry(width = 5)
ad_value6.grid(column=0, row=8)
Label(window, text="неудовлетв., кол-во:").grid(row=9, column=0, sticky=W)
ad_value7:Entry = ttk.Entry(width = 5)
ad_value7.grid(column=0, row=9)
Label(window, text="Диаметр, см:").grid(row=1, column=1, sticky=W)
ad_value8:Entry = ttk.Entry(width = 5)
ad_value8.grid(column=1, row=2)

wb = openpyxl.load_workbook('1.xlsx')
ws = wb['1']

from openpyxl.drawing.image import Image

# создаем объект Image из файла изображения
img = Image('Screenshot_1.png')

# получаем координаты из anchor объекта
##_, _, _, _ = img.anchor


##print(f"Координаты изображения: ({left}, {top})")

wb.save('1.xlsx')
wb.close()

window.mainloop()




